import io
import csv
from collections import defaultdict

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, update, or_
from pydantic import BaseModel
from typing import Optional, List

from app.db.session import get_db
from app.models.models import (
    Order, OrderLine, User, PodRegistry,
    PodDocument, PackingSlipDocument, InvoiceDocument, EmailRequest,
)
from app.core.security import get_current_user, require_admin

router = APIRouter()


class OrderLineIn(BaseModel):
    line_number: int
    material_number: Optional[str] = None
    material_description: Optional[str] = None
    lot_number: Optional[str] = None
    quantity: Optional[float] = None
    unit_of_measure: Optional[str] = None
    tracking_number: Optional[str] = None
    carrier: Optional[str] = "UPS"


class OrderIn(BaseModel):
    customer_order_number: str
    my_delivery_number: Optional[str] = None
    warehouse_delivery_number: Optional[str] = None
    sales_order_number: Optional[str] = None
    invoice_number: Optional[str] = None
    customer_name: Optional[str] = None
    customer_email: Optional[str] = None
    lines: Optional[List[OrderLineIn]] = []


async def _upsert_pod_registry(db: AsyncSession, delivery_number: str, order_id, customer_po: str = None) -> None:
    """Create a pending pod_registry entry if one doesn't already exist for this delivery number."""
    r = await db.execute(
        select(PodRegistry).where(PodRegistry.delivery_number == delivery_number)
    )
    existing = r.scalar_one_or_none()
    if not existing:
        db.add(PodRegistry(
            delivery_number=delivery_number,
            order_id=order_id,
            customer_po=customer_po,
            status='pending',
        ))
    elif customer_po and not existing.customer_po:
        existing.customer_po = customer_po
    await db.flush()


@router.post("")
async def create_order(data: OrderIn, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    order = Order(**data.model_dump(exclude={"lines"}))
    db.add(order)
    await db.flush()
    for line in (data.lines or []):
        db.add(OrderLine(order_id=order.id, **line.model_dump()))
    if order.my_delivery_number:
        await _upsert_pod_registry(db, order.my_delivery_number, order.id, order.customer_order_number)
    await db.commit()
    # Trigger Power Automate desktop flows if configured (fire-and-forget)
    from app.core.tasks import trigger_power_automate_task
    trigger_power_automate_task.delay([str(order.id)])
    return {"id": str(order.id), "customer_order_number": order.customer_order_number}


@router.get("/template")
async def download_order_template(_: User = Depends(get_current_user)):
    """Return a formatted .xlsx import template for bulk order upload."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import openpyxl.utils

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = [
        'customer_order_number', 'my_delivery_number', 'warehouse_delivery_number',
        'sales_order_number', 'invoice_number', 'customer_name', 'customer_email',
        'line_number', 'material_number', 'material_description', 'lot_number',
        'quantity', 'unit_of_measure', 'tracking_number', 'carrier',
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Example data row
    ws.append([
        'PO-1234', 'DEL-2024-0001', 'WH-001', 'SO-5678', 'INV-9012',
        'ACME Corp', 'acme@example.com', 1, 'MAT-001', 'Widget A',
        'LOT-001', 100, 'EA', '1Z999AA10123456784', 'UPS',
    ])

    # Column widths
    col_widths = [24, 20, 24, 18, 16, 22, 28, 12, 16, 28, 14, 10, 14, 24, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="order_import_template.xlsx"'},
    )


@router.post("/import")
async def import_orders(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """
    Bulk import orders from a .xlsx or .csv file.

    Uniqueness rules:
      - Order identity:  my_delivery_number  (one Order per delivery number)
      - OrderLine identity: (my_delivery_number, material_number, lot_number)

    On import:
      - Row without my_delivery_number → error.
      - Order not found → create new Order; line added → created.
      - Order found + matching line (same NDC + lot) with same quantity → skipped.
      - Order found + matching line with different quantity → quantity updated (latest-wins) → updated.
      - Order found + no matching line → new line added → created.
    """
    content = await file.read()
    filename = (file.filename or '').lower()

    rows: list = []
    if filename.endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        rows = [dict(r) for r in reader]
    elif filename.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
    else:
        raise HTTPException(400, "Only .xlsx or .csv files are supported")

    if not rows:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": ["No data rows found in file"]}

    def _s(v):
        """Return stripped string or None."""
        s = str(v or '').strip()
        return s or None

    def _qty(v):
        try:
            return float(v) if v not in (None, '') else None
        except (ValueError, TypeError):
            return None

    created = 0
    updated = 0
    skipped = 0
    errors: list = []
    new_order_ids: list = []

    # Cache Orders we've already looked up / created during this import,
    # keyed by my_delivery_number, to avoid re-querying per row.
    order_cache: dict = {}

    for idx, row in enumerate(rows, start=2):  # row 1 is header
        mdn = _s(row.get('my_delivery_number'))
        if not mdn:
            errors.append(f"row {idx}: missing my_delivery_number")
            continue
        try:
            async with db.begin_nested():
                order = order_cache.get(mdn)
                if order is None:
                    r = await db.execute(
                        select(Order).where(Order.my_delivery_number == mdn)
                    )
                    order = r.scalars().first()
                    if order is None:
                        order = Order(
                            customer_order_number=_s(row.get('customer_order_number')) or mdn,
                            my_delivery_number=mdn,
                            warehouse_delivery_number=_s(row.get('warehouse_delivery_number')),
                            sales_order_number=_s(row.get('sales_order_number')),
                            invoice_number=_s(row.get('invoice_number')),
                            customer_name=_s(row.get('customer_name')),
                            customer_email=_s(row.get('customer_email')),
                        )
                        db.add(order)
                        await db.flush()
                        await _upsert_pod_registry(db, mdn, order.id, order.customer_order_number)
                        new_order_ids.append(str(order.id))
                    order_cache[mdn] = order

                material = _s(row.get('material_number'))
                lot = _s(row.get('lot_number'))
                new_qty = _qty(row.get('quantity'))

                # Find existing line by (order_id, material_number, lot_number).
                # Use .first() to tolerate any historical duplicates.
                line_q = await db.execute(
                    select(OrderLine).where(
                        OrderLine.order_id == order.id,
                        OrderLine.material_number == material,
                        OrderLine.lot_number == lot,
                    )
                )
                existing_line = line_q.scalars().first()

                if existing_line is not None:
                    old_qty = float(existing_line.quantity) if existing_line.quantity is not None else None
                    if old_qty == new_qty:
                        skipped += 1
                    else:
                        existing_line.quantity = new_qty
                        updated += 1
                else:
                    try:
                        line_num = int(row.get('line_number')) if row.get('line_number') not in (None, '') else None
                    except (ValueError, TypeError):
                        line_num = None
                    if line_num is None:
                        # Fall back to (count of existing lines + 1) so numbering stays stable.
                        cnt_q = await db.execute(
                            select(OrderLine).where(OrderLine.order_id == order.id)
                        )
                        line_num = len(cnt_q.scalars().all()) + 1
                    db.add(OrderLine(
                        order_id=order.id,
                        line_number=line_num,
                        material_number=material,
                        material_description=_s(row.get('material_description')),
                        lot_number=lot,
                        quantity=new_qty,
                        unit_of_measure=_s(row.get('unit_of_measure')),
                        tracking_number=_s(row.get('tracking_number')),
                        carrier=_s(row.get('carrier')) or 'UPS',
                    ))
                    created += 1
        except Exception as e:
            errors.append(f"row {idx} (delivery {mdn}): {str(e)}")

    await db.commit()
    # Trigger Power Automate desktop flows if configured (fire-and-forget) — only for newly created orders
    if new_order_ids:
        from app.core.tasks import trigger_power_automate_task
        trigger_power_automate_task.delay(new_order_ids)
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


@router.delete("/{order_id}")
async def delete_order(
    order_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """
    Delete an order and cascade cleanly:
      - order_lines: auto-deleted by DB CASCADE
      - pod_registry (status=pending): deleted (only existed because the order did)
      - pod_registry (any other status): order_id nulled (real POD may be attached)
      - pod_documents, packing_slip_documents, invoice_documents: order_id nulled
      - email_requests: order_id nulled
    """
    order = await db.get(Order, order_id)
    if not order:
        raise HTTPException(404, "Order not found")

    con = order.customer_order_number

    # ── 1. Handle pod_registry ────────────────────────────────
    reg_result = await db.execute(
        select(PodRegistry).where(PodRegistry.order_id == order.id)
    )
    for reg in reg_result.scalars().all():
        if str(reg.status) == 'pending':
            await db.delete(reg)
        else:
            reg.order_id = None

    # ── 2. Null out order_id on document tables ───────────────
    for model in (PodDocument, PackingSlipDocument, InvoiceDocument):
        await db.execute(
            update(model).where(model.order_id == order.id).values(order_id=None)
        )

    # ── 3. Null out order_id on email_requests ────────────────
    await db.execute(
        update(EmailRequest).where(EmailRequest.order_id == order.id).values(order_id=None)
    )

    # ── 4. Delete the order (order_lines cascade via DB) ──────
    await db.delete(order)
    await db.commit()

    return {"deleted": True, "customer_order_number": con}


@router.get("")
async def list_orders(search: Optional[str] = None, limit: int = 50, offset: int = 0, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    stmt = select(Order).order_by(desc(Order.created_at))
    if search:
        term = f"%{search}%"
        stmt = stmt.where(or_(
            Order.customer_order_number.ilike(term),
            Order.my_delivery_number.ilike(term),
            Order.warehouse_delivery_number.ilike(term),
            Order.customer_name.ilike(term),
            Order.sales_order_number.ilike(term),
            Order.invoice_number.ilike(term),
        ))
        limit = 500
    result = await db.execute(stmt.limit(limit).offset(offset))
    orders = result.scalars().all()
    if not orders:
        return []
    order_ids = [o.id for o in orders]
    lines_result = await db.execute(
        select(OrderLine).where(OrderLine.order_id.in_(order_ids))
    )
    lines_by_order: dict = defaultdict(list)
    for line in lines_result.scalars().all():
        lines_by_order[line.order_id].append(line)
    return [_order_out(o, lines_by_order[o.id]) for o in orders]


@router.get("/{order_id}")
async def get_order(order_id: str, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    o = await db.get(Order, order_id)
    if not o:
        raise HTTPException(404)
    lines_result = await db.execute(select(OrderLine).where(OrderLine.order_id == o.id))
    return _order_out(o, lines_result.scalars().all())


@router.put("/{order_id}")
async def update_order(order_id: str, data: OrderIn, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    o = await db.get(Order, order_id)
    if not o:
        raise HTTPException(404)
    for k, v in data.model_dump(exclude={"lines"}).items():
        if v is not None:
            setattr(o, k, v)
    await db.commit()
    return {"id": str(o.id)}


def _order_out(o: Order, lines):
    return {
        "id": str(o.id),
        "customer_order_number": o.customer_order_number,
        "my_delivery_number": o.my_delivery_number,
        "warehouse_delivery_number": o.warehouse_delivery_number,
        "sales_order_number": o.sales_order_number,
        "invoice_number": o.invoice_number,
        "customer_name": o.customer_name,
        "customer_email": o.customer_email,
        "status": o.status,
        "created_at": o.created_at.isoformat() if o.created_at else None,
        "lines": [{
            "id": str(l.id),
            "line_number": l.line_number,
            "material_number": l.material_number,
            "material_description": l.material_description,
            "lot_number": l.lot_number,
            "quantity": float(l.quantity) if l.quantity else None,
            "unit_of_measure": l.unit_of_measure,
            "tracking_number": l.tracking_number,
            "carrier": l.carrier,
        } for l in lines],
    }
