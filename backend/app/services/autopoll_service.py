"""
autopoll_service.py
Polls a configured directory for xlsx/csv order import files on a schedule.

Flow per cycle:
  1. Read autopoll_* keys from system_config
  2. Scan autopoll_path for *.xlsx / *.csv (skipping processed/ subdirectory)
  3. Parse each file, import unique orders (dedup by my_delivery_number first)
  4. Move processed files to {autopoll_path}/processed/{timestamp}_{filename}
  5. Update autopoll_last_run and autopoll_last_result in system_config
  6. If any orders created: dispatch poll_ftp_task + scan_order_documents_task
"""
import io
import csv
import json
import logging
import os
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.models import Order, OrderLine, PodRegistry, SystemConfig

logger = logging.getLogger(__name__)

# Keys used in system_config
_KEY_ENABLED  = "autopoll_enabled"
_KEY_PATH     = "autopoll_path"
_KEY_FREQ     = "autopoll_frequency_minutes"
_KEY_LAST_RUN = "autopoll_last_run"
_KEY_LAST_RES = "autopoll_last_result"

DEFAULT_PATH  = "./storage/order_import"
DEFAULT_FREQ  = 5


async def get_autopoll_config(db: AsyncSession) -> dict:
    keys = [_KEY_ENABLED, _KEY_PATH, _KEY_FREQ, _KEY_LAST_RUN, _KEY_LAST_RES]
    result = await db.execute(select(SystemConfig).where(SystemConfig.key.in_(keys)))
    cfg = {r.key: r.value for r in result.scalars().all()}
    return {
        "enabled":            cfg.get(_KEY_ENABLED, "false").lower() == "true",
        "path":               cfg.get(_KEY_PATH, DEFAULT_PATH) or DEFAULT_PATH,
        "frequency_minutes":  int(cfg.get(_KEY_FREQ, str(DEFAULT_FREQ)) or DEFAULT_FREQ),
        "last_run":           cfg.get(_KEY_LAST_RUN),
        "last_result":        _safe_json(cfg.get(_KEY_LAST_RES)),
    }


async def check_and_run_autopoll(db: AsyncSession) -> Optional[dict]:
    """
    Called from the background loop every 30 s.
    Runs the actual poll only when enabled AND the configured interval has elapsed.
    Returns the result dict, or None if skipped.
    """
    cfg = await get_autopoll_config(db)
    if not cfg["enabled"]:
        return None

    now = datetime.now()
    last_run_str = cfg["last_run"]
    if last_run_str:
        try:
            last_run = datetime.fromisoformat(last_run_str)
            if last_run.tzinfo is not None:
                last_run = last_run.replace(tzinfo=None)
            elapsed = now - last_run
            if elapsed < timedelta(minutes=cfg["frequency_minutes"]):
                return None
        except ValueError:
            pass  # bad timestamp → run now

    return await run_autopoll(db)


async def run_autopoll(db: AsyncSession) -> dict:
    """
    Full autopoll cycle. Scans the configured directory, imports unique orders,
    moves processed files, updates system_config, and triggers downstream tasks.
    """
    cfg = await get_autopoll_config(db)
    poll_path = Path(cfg["path"])

    result = {
        "files_found":     0,
        "files_processed": 0,
        "created":         0,
        "skipped":         0,
        "errors":          [],
        "new_order_ids":   [],
    }

    # Ensure the directory exists
    try:
        poll_path.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        result["errors"].append(f"Cannot access autopoll path {poll_path}: {e}")
        await _save_result(db, result)
        return result

    processed_dir = poll_path / "processed"
    try:
        processed_dir.mkdir(exist_ok=True)
    except Exception as e:
        result["errors"].append(f"Cannot create processed/ subdirectory: {e}")
        await _save_result(db, result)
        return result

    # Find eligible files (exclude the processed/ subdirectory itself)
    eligible = [
        f for f in poll_path.iterdir()
        if f.is_file() and f.suffix.lower() in (".xlsx", ".csv")
    ]
    result["files_found"] = len(eligible)

    for filepath in eligible:
        file_result = await _process_file(db, filepath)
        result["created"]  += file_result["created"]
        result["skipped"]  += file_result["skipped"]
        result["errors"]   += file_result["errors"]
        result["new_order_ids"] += file_result["new_order_ids"]

        # Move to processed/ regardless of whether rows were created or skipped
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = processed_dir / f"{timestamp}_{filepath.name}"
        try:
            filepath.rename(dest)
            result["files_processed"] += 1
            logger.info("Autopoll: moved %s → %s", filepath.name, dest.name)
        except Exception as e:
            result["errors"].append(f"Could not move {filepath.name} to processed/: {e}")

    # Persist result and run timestamp
    await _save_result(db, result)

    # Trigger downstream document-finding tasks for newly created orders
    new_ids = result["new_order_ids"]
    if new_ids:
        _dispatch_downstream(new_ids)

    return result


async def _process_file(db: AsyncSession, filepath: Path) -> dict:
    """Parse one xlsx/csv file and import unique orders."""
    out = {"created": 0, "skipped": 0, "errors": [], "new_order_ids": []}

    try:
        content = filepath.read_bytes()
    except Exception as e:
        out["errors"].append(f"{filepath.name}: cannot read file — {e}")
        return out

    rows: list = []
    try:
        if filepath.suffix.lower() == ".csv":
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows.append(dict(zip(headers, row)))
    except Exception as e:
        out["errors"].append(f"{filepath.name}: parse error — {e}")
        return out

    if not rows:
        out["errors"].append(f"{filepath.name}: no data rows found")
        return out

    def _s(v):
        s = str(v or "").strip()
        return s or None

    # ── Group by my_delivery_number (primary key) then customer_order_number ──
    # Rows that have a delivery number → keyed by delivery number
    # Rows with no delivery number → keyed by customer_order_number (fallback)
    by_delivery: dict = defaultdict(list)   # delivery_number → rows
    by_con: dict      = defaultdict(list)   # customer_order_number → rows (no dn)
    ungrouped         = 0

    for row in rows:
        dn  = _s(row.get("my_delivery_number"))
        con = _s(row.get("customer_order_number"))
        if dn:
            by_delivery[dn].append(row)
        elif con:
            by_con[con].append(row)
        else:
            ungrouped += 1

    if ungrouped:
        out["errors"].append(
            f"{filepath.name}: {ungrouped} row(s) skipped — no delivery_number or customer_order_number"
        )

    # ── Process delivery-number-keyed groups ──────────────────────────────────
    for dn, group_rows in by_delivery.items():
        try:
            async with db.begin_nested():
                # Primary dedup key: my_delivery_number
                existing = await db.execute(
                    select(Order).where(Order.my_delivery_number == dn)
                )
                if existing.scalar_one_or_none():
                    out["skipped"] += 1
                    continue

                first = group_rows[0]
                order = Order(
                    customer_order_number=_s(first.get("customer_order_number")) or dn,
                    my_delivery_number=dn,
                    warehouse_delivery_number=_s(first.get("warehouse_delivery_number")),
                    sales_order_number=_s(first.get("sales_order_number")),
                    invoice_number=_s(first.get("invoice_number")),
                    customer_name=_s(first.get("customer_name")),
                    customer_email=_s(first.get("customer_email")),
                )
                db.add(order)
                await db.flush()

                for i, row in enumerate(group_rows):
                    db.add(OrderLine(
                        order_id=order.id,
                        line_number=_safe_int(row.get("line_number"), i + 1),
                        material_number=_s(row.get("material_number")),
                        material_description=_s(row.get("material_description")),
                        lot_number=_s(row.get("lot_number")),
                        quantity=_safe_float(row.get("quantity")),
                        unit_of_measure=_s(row.get("unit_of_measure")),
                        tracking_number=_s(row.get("tracking_number")),
                        carrier=_s(row.get("carrier")) or "UPS",
                    ))

                await _upsert_pod_registry(db, dn, order.id, order.customer_order_number)
                out["created"] += 1
                out["new_order_ids"].append(str(order.id))

        except Exception as e:
            out["errors"].append(f"{filepath.name} [{dn}]: {e}")

    # ── Process fallback customer_order_number-keyed groups ───────────────────
    for con, group_rows in by_con.items():
        try:
            async with db.begin_nested():
                existing = await db.execute(
                    select(Order).where(Order.customer_order_number == con)
                )
                if existing.scalar_one_or_none():
                    out["skipped"] += 1
                    continue

                first = group_rows[0]
                order = Order(
                    customer_order_number=con,
                    my_delivery_number=None,
                    warehouse_delivery_number=_s(first.get("warehouse_delivery_number")),
                    sales_order_number=_s(first.get("sales_order_number")),
                    invoice_number=_s(first.get("invoice_number")),
                    customer_name=_s(first.get("customer_name")),
                    customer_email=_s(first.get("customer_email")),
                )
                db.add(order)
                await db.flush()

                for i, row in enumerate(group_rows):
                    db.add(OrderLine(
                        order_id=order.id,
                        line_number=_safe_int(row.get("line_number"), i + 1),
                        material_number=_s(row.get("material_number")),
                        material_description=_s(row.get("material_description")),
                        lot_number=_s(row.get("lot_number")),
                        quantity=_safe_float(row.get("quantity")),
                        unit_of_measure=_s(row.get("unit_of_measure")),
                        tracking_number=_s(row.get("tracking_number")),
                        carrier=_s(row.get("carrier")) or "UPS",
                    ))

                out["created"] += 1
                out["new_order_ids"].append(str(order.id))

        except Exception as e:
            out["errors"].append(f"{filepath.name} [{con}]: {e}")

    await db.commit()
    return out


async def _upsert_pod_registry(db: AsyncSession, delivery_number: str, order_id, customer_po: str = None) -> None:
    r = await db.execute(
        select(PodRegistry).where(PodRegistry.delivery_number == delivery_number)
    )
    existing = r.scalar_one_or_none()
    if not existing:
        db.add(PodRegistry(
            delivery_number=delivery_number,
            order_id=order_id,
            customer_po=customer_po,
            status="pending",
        ))
    elif customer_po and not existing.customer_po:
        existing.customer_po = customer_po
    await db.flush()


async def _save_result(db: AsyncSession, result: dict) -> None:
    now_iso = datetime.now().isoformat()
    # Store a copy without the order IDs list (too large for config value)
    summary = {k: v for k, v in result.items() if k != "new_order_ids"}
    for key, value in [
        (_KEY_LAST_RUN, now_iso),
        (_KEY_LAST_RES, json.dumps(summary)),
    ]:
        cfg = await db.get(SystemConfig, key)
        if not cfg:
            cfg = SystemConfig(key=key)
            db.add(cfg)
        cfg.value = value
        cfg.updated_at = datetime.now()
    await db.commit()


def _dispatch_downstream(new_order_ids: list) -> None:
    """Dispatch Celery tasks to find PODs, packing slips, and invoices for new orders."""
    try:
        from app.core.tasks import poll_ftp_task, scan_order_documents_task, trigger_power_automate_task
        poll_ftp_task.delay()
        scan_order_documents_task.delay(new_order_ids)
        trigger_power_automate_task.delay(new_order_ids)
        logger.info("Autopoll: dispatched FTP poll + document scan + Power Automate for %d new order(s)", len(new_order_ids))
    except Exception as e:
        logger.error("Autopoll: failed to dispatch downstream tasks: %s", e)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_int(v, default: int = 1) -> int:
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _safe_float(v) -> Optional[float]:
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _safe_json(s: Optional[str]) -> Optional[dict]:
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        return None
